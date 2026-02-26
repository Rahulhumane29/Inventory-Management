from django.http import HttpResponse
from rest_framework.decorators import api_view
from .models import Inventory_items
from .serializers import Itemserializers
from rest_framework.response import Response
from rest_framework import status
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from io import BytesIO
from reportlab.lib.pagesizes import letter
import openpyxl
from openpyxl.styles import Font, PatternFill
from .filters import Itemfilters


# Create your views here.


@api_view(['GET', 'POST'])
def items(request):
    if request.method == 'GET':
        queryset = Inventory_items.objects.all()
        filtered_items = Itemfilters(request.GET, queryset=queryset)
        serializer = Itemserializers(filtered_items.qs, many=True)

        return Response(serializer.data) 

    elif request.method == 'POST':
        if isinstance(request.data, list):
            serializer = Itemserializers(data=request.data, many=True)
        else:
            serializer = Itemserializers(data=request.data)       
            
            
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Inventory items edit
@api_view(['GET', 'PUT', 'DELETE'])
def items_details(request, pk):
    try:
        items = Inventory_items.objects.get(pk=pk)
    except Inventory_items.DoesNotExist:
        return Response({'error': 'Inventory Items not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = Itemserializers(items, context={'request': request})
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = Itemserializers(items, data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        items.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
def export_items_pdf(request):

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    data = [['ID', 'Name', 'Quantity', 'Price', 'Created At']]

    items = Inventory_items.objects.all()

    for item in items:
        data.append([
            str(item.id),
            str(item.name),
            str(item.quantity),
            str(item.price),
            item.created_at.strftime("%d-%m-%Y %H:%M")
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="items_report.pdf"'
    response.write(pdf)

    return response


@api_view(['GET'])
def export_items_excel(request):
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    headers = ['ID', 'Name', 'Quantity', 'Price ($)', 'Created At']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    ws.append(headers)
    
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    items = Inventory_items.objects.all()

    for item in items:
        ws.append([
            item.id,
            item.name,
            item.quantity,
            float(item.price),
            item.created_at.strftime("%Y-%m-%d %H:%M")
        ])
        
        # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename="inventory_items.xlsx"'

    wb.save(response)

    return response

    
    